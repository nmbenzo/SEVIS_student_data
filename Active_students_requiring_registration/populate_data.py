import os
os.getcwd()
import openpyxl

wb2 = openpyxl.load_workbook('/Users/nbenzschawel/Downloads/Fall 201840 - Registration.xlsx')

sheet = wb2.worksheets[4]

ws2 = wb2.active


def match_SEVISID(campusID_SEVISID):
    """
    match_SEVISID compares SEVISIDs from two separate workbooks(ws, ws2) 
    and then populates a blank column with the appropriate BannerID(campusID)
    """
    sheet.insert_cols(1)
    title = sheet.cell(row=1, column=1)
    title.value = 'campusID'
    for rowNum in range(2, sheet.max_row):
        sevis_ID = sheet.cell(row=rowNum, column=2).value
        for x in campusID_SEVISID:
            if x == sevis_ID:
                sheet.cell(row=rowNum, column=1).value = campusID_SEVISID[sevis_ID]

    wb2.save('/Users/nbenzschawel/Downloads/Fall 201840 - Registration.xlsx')


def match_major_data(SEVISID_major):
    """match_major_data compares SEVISIDs against major data and then populates
    a blank column with the appropriate major data when a SEVISID match is found.
    """
    sheet.insert_cols(1)
    title = sheet.cell(row=1, column=1)
    title.value = 'Major'
    for rowNum in range(2, sheet.max_row):
        sevis_ID = sheet.cell(row=rowNum, column=3).value
        for x in SEVISID_major:
            if x == sevis_ID:
                sheet.cell(row=rowNum, column=1).value = SEVISID_major[sevis_ID]

    wb2.save('/Users/nbenzschawel/Downloads/Fall 201840 - Registration.xlsx')


def match_advisor(advisor_major_ug, advisor_major_gr):
    """
    match_advisor checks the majors in a column of workbook(ws) and
    then matches them with the advisor in a dictionary from the module: data
    """
    sheet.insert_cols(1)
    title = sheet.cell(row=1, column=1)
    title.value = 'Advisor'
    for rowNum in range(2, sheet.max_row):
        major = sheet.cell(row=rowNum, column=2).value
        if major in advisor_major_ug:
            sheet.cell(row=rowNum, column=1).value = advisor_major_ug[major]
        if major in advisor_major_gr:
            sheet.cell(row=rowNum, column=1).value = advisor_major_gr[major]

    wb2.save('/Users/nbenzschawel/Downloads/Fall 201840 - Registration.xlsx')


