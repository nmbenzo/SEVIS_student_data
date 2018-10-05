import os
os.getcwd()
import openpyxl
from SEVIS_grad_data.build_grad_data import ws, campusID_SEVISID, campusID_work_auth, campusID_end_date, campusID_workend, campusID_emails
from major_advisor_data import master_deg_adv, bachelor_deg_adv

wb2 = openpyxl.load_workbook('/Users/nbenzschawel/Downloads/201840_graduated_students_test.xlsx')
wb3Live = openpyxl.load_workbook('/Users/nbenzschawel/Desktop/201820_201830_graduated_students.xlsx')

sheet2 = wb2.worksheets[0]
sheet3 = wb3Live.worksheets[0]

ws2 = wb2.active
ws3 = wb3Live.active


def add_advisor(master_deg_adv, bachelor_deg_adv):
    """match_advisor checks the majors in a column of workbook(ws) and
    then matches them with the advisor in a dictionary from the module:
    build_grad_data
    """
    sheet2.insert_cols(1)
    title = ws2.cell(row=1, column=1)
    title.value = 'Advisor'
    for rowNum in range(2, ws2.max_row):
        major = ws2.cell(row=rowNum, column=6).value
        if major in master_deg_adv:
            ws2.cell(row=rowNum, column=1).value = master_deg_adv[major]
        if major in bachelor_deg_adv:
            ws2.cell(row=rowNum, column=1).value = bachelor_deg_adv[major]

    wb2.save('/Users/nbenzschawel/Downloads/201840_graduated_students_test.xlsx')


def match_SEVISID(campusID_SEVISID):
    """Builds a column of SEVISIDs for students based on the
    campusID_SEVISID dictionary
    """
    sheet2.insert_cols(1)
    title = ws2.cell(row=1, column=1)
    title.value = 'SEVISID'
    for rowNum in range(2, ws2.max_row):
        campusID = ws2.cell(row=rowNum, column=3).value
        for x in campusID_SEVISID:
            if x == campusID:
                ws2.cell(row=rowNum, column=1).value = campusID_SEVISID[campusID]

    wb2.save('/Users/nbenzschawel/Downloads/201840_graduated_students_test.xlsx')


def add_work_type(campusID_work_auth):
    """Builds an empty column. Loops through a range of campusIDs and then
    compares these against campusIDs in campusID_work_auth dictionary. When it
    finds a matched campusID in the campusID_work_auth dictionary it populates
    the empty column with the appropriate work authorization type
    """
    sheet2.insert_cols(1)
    title = ws2.cell(row=1, column=1)
    title.value = 'Work Authorization Type'
    for rowNum in range(2, ws2.max_row):
        campusID = ws2.cell(row=rowNum, column=4).value
        for x in campusID_work_auth:
            if x == campusID:
                ws2.cell(row=rowNum, column=1).value = campusID_work_auth[campusID]

    wb2.save('/Users/nbenzschawel/Downloads/201840_graduated_students_test.xlsx')


def add_work_enddate(campusID_workend):
    """Builds an empty column. Loops through a range of campusIDs and then
    compares these against campusIDs in campusID_workend dictionary. When it
    finds a matched campusID in the campusID_workend dictionary it populates
    the empty column with the appropriate work authorization end date.
    """
    sheet2.insert_cols(1)
    title = ws2.cell(row=1, column=1)
    title.value = 'Work Authorization End Date'
    for rowNum in range(2, ws2.max_row):
        campusID = ws2.cell(row=rowNum, column=5).value
        for i in campusID_workend:
            if i == campusID:
                ws2.cell(row=rowNum, column=1).value = campusID_workend[campusID]

    wb2.save('/Users/nbenzschawel/Downloads/201840_graduated_students_test.xlsx')


def add_profile_enddate(campusID_end_date):
    """Builds an empty column. Loops through a range of campusIDs and then
    compares these against campusIDs in campusID_end_date dictionary. When it
    finds a matched campusID in the campusID_end_date dictionary it populates
    the empty column with the appropriate profile end date.
    """
    sheet2.insert_cols(1)
    title = ws2.cell(row=1, column=1)
    title.value = 'Profile End Date'
    for rowNum in range(2, ws2.max_row):
        campusID = ws2.cell(row=rowNum, column=6).value
        for i in campusID_end_date:
            if i == campusID:
                ws2.cell(row=rowNum, column=1).value = campusID_end_date[campusID]

    wb2.save('/Users/nbenzschawel/Downloads/201840_graduated_students_test.xlsx')



def add_emails(campusID_emails):
    sheet3.insert_cols(1)
    title = ws3.cell(row=1, column=1)
    title.value = 'Emails'
    for rowNum in range(2, ws2.max_row):
        campusID = ws3.cell(row=rowNum, column=6).value
        for i in campusID_emails:
            if i == campusID:
                ws3.cell(row=rowNum, column=1).value = campusID_emails[campusID]

    wb3Live.save('/Users/nbenzschawel/Desktop/201820_201830_graduated_students.xlsx')



add_advisor(master_deg_adv, bachelor_deg_adv)
match_SEVISID(campusID_SEVISID)
add_work_type(campusID_work_auth)
add_work_enddate(campusID_workend)
add_profile_enddate(campusID_end_date)
add_emails(campusID_emails)