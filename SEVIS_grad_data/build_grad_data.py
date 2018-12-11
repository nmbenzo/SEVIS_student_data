import os
os.getcwd()
import openpyxl
from Handlers.file_imports import wb1_grad, sheet_grad, ws_grad

"""
This file builds dictionaries containing the following data: 
campusIDs : SEVISIDs
campusIDs : Work Authorization Types
campusIDs : Work Authorization End Dates
campusIDs : Profile End Dates
campusIDs : Student Emails

"""


campusID = []

for rowNum in range(2, ws_grad.max_row):
    id = ws_grad.cell(row=rowNum, column=1).value
    campusID.append(id)


SEVISID = []

for rowNum in range(2, ws_grad.max_row):
    id = ws_grad.cell(row=rowNum, column=5).value
    SEVISID.append(id)


campusID_SEVISID = {}

for i in range(len(campusID)): # verbose dict building
    campusID_SEVISID[campusID[i]] = SEVISID[i]


work_auth_type = []

for rowNum in range(2, ws_grad.max_row):
    type = ws_grad.cell(row=rowNum, column=31).value # need to check column in spreadsheet
    work_auth_type.append(type)

campusID_work_auth = dict(zip(campusID, work_auth_type))


work_auth_enddate = []

for x in range(2, ws_grad.max_row):
    work_end = ws_grad.cell(row=x, column=33).value # need to check column in spreadsheet
    work_auth_enddate.append(work_end)

campusID_workend = dict(zip(campusID, work_auth_enddate))


profile_end_date = []

for rowNum in range(2, ws_grad.max_row):
    end_date = ws_grad.cell(row=rowNum, column=10).value # need to check column in spreadsheet
    profile_end_date.append(end_date)

campusID_end_date = dict(zip(campusID, profile_end_date))

emails = []
email_campID = []

for x in range(2, ws_grad.max_row):
    email = ws_grad.cell(row=x, column=30).value # need to check column in spreadsheet
    emails.append(email)

for x in range(2, ws_grad.max_row):
    campusID = ws_grad.cell(row=x, column=1).value
    email_campID.append(campusID)

campusID_emails = dict(zip(email_campID, emails))




