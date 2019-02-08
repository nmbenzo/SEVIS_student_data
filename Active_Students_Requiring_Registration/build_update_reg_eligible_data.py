import openpyxl
from Handlers.major_advisor_data import *
from Active_Students_Requiring_Registration.populate_active_data import *
from Active_Students_Requiring_Registration.build_data import *
from Active_Students_Requiring_Registration.sort_active_data import *
from tqdm import trange

"""
This file can be run when the office is ready to complete the register 
continuing students. In order for this file function properly, you must have 
already downloaded both the UG and GR Tracking_OK_to_Register.xlsx files and 
the SEVIS-Active_Students Requiring Registration File.
"""

""" ISSM ACTIVE Students Eligible for Registration Data"""

ACTIVE_UG_eligible_for_registration = \
    '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/' \
    'SEVIS_Registration_Tracking-Continuing_UG_Students-OK_to_Register.xlsx'

ACTIVE_GR_eligible_for_registration = \
    '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/' \
    'SEVIS_Registration_Tracking-Continuing_Grad_Students-OK_to_Register.xlsx'

ACTIVE_GR_reg_wb = openpyxl.load_workbook(ACTIVE_GR_eligible_for_registration)
gr_active_sheet = ACTIVE_GR_reg_wb.worksheets[0]

ACTIVE_UG_reg_wb = openpyxl.load_workbook(ACTIVE_UG_eligible_for_registration)
ug_active_sheet = ACTIVE_UG_reg_wb.worksheets[0]

Reg_Eligible_Students_FINAL = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/' \
                              'Spring/Final Workbooks/REG_ELIGIBLE_FINAL.xlsx'

REG_ELIGIBLE_wb = openpyxl.Workbook()

try:
    Reg_eligible_students = openpyxl.load_workbook(Reg_Eligible_Students_FINAL)
    Eligible_sheet = Reg_eligible_students.worksheets[0]
except FileNotFoundError:
    REG_ELIGIBLE_wb.save(Reg_Eligible_Students_FINAL)
    Reg_eligible_students = openpyxl.load_workbook(Reg_Eligible_Students_FINAL)
    Eligible_sheet = Reg_eligible_students.worksheets[0]

reg_ug_row_max = ug_active_sheet.max_row
reg_ug_col_max = ug_active_sheet.max_column
reg_gr_row_max = gr_active_sheet.max_row
reg_gr_col_max = gr_active_sheet.max_column


def create_Reg_Eligible_Data():
    """
    Creates a final workbook of all UG and GR students that are eligible for
    registration based on reports run in ISSM
    """
    if ug_active_sheet.cell(row=1, column=1).value == 'Campus Id':
        selectedRange = copy_new_Range(1,2, reg_gr_col_max, reg_gr_row_max, gr_active_sheet)
        paste_new_Range(1, (reg_ug_row_max+1), reg_gr_col_max,
                        (reg_ug_row_max + reg_gr_row_max-1), ug_active_sheet, selectedRange)
        # In order to save to the Reg_Eligible workbook, need to use
        # ACTIVE_UG_reb_work as your saving source.
        ACTIVE_UG_reg_wb.save(Reg_Eligible_Students_FINAL)

def build_campusID_list():
    df = pd.ExcelFile(Reg_Eligible_Students_FINAL).parse('SEVIS_Registration_Tracking-Con')
    Reg_CampusID = df['Campus Id'].tolist()
    return Reg_CampusID


def update_registration_status(create_Reg_Eligible_Data):
    """update_registration_status checks the """
    ACTIVE_students_sheet.insert_cols(1)
    title = ACTIVE_students_sheet.cell(row=1, column=1)
    title.value = 'Registration Status'
    for rowNum in trange(2, ACTIVE_students_sheet.max_row):
        campusID = ACTIVE_students_sheet.cell(row=rowNum, column=5).value
        for x in create_Reg_Eligible_Data:
            if x == campusID:
                ACTIVE_students_sheet.cell(row=rowNum, column=1).value = \
                'Registration Submitted'

    ACTIVE_students.save(ACTIVE_students_FINAL)


print('\nCreating data now...')
time.sleep(0.5)
print('\nBeginning data aggregation for ACTIVE Status students...')
create_active_Student_Data()
time.sleep(1.3)
active_match_SEVISID(active_campusID_SEVISID)
time.sleep(0.3)
match_units(active_SEVISID_units)
time.sleep(0.3)
match_major_data(active_SEVISID_major)
time.sleep(0.3)
match_advisor(advisor_major_ug, advisor_major_gr)

time.sleep(0.5)
create_Reg_Eligible_Data()
time.sleep(0.5)
build_campusID_list()
time.sleep(1)
update_registration_status(build_campusID_list())
time.sleep(1)
print('Updated the registration status for students eligible for registration')
time.sleep(0.5)
sort_active_data()
print(
    '\n*** Ran all functions for building notes for SEVIS Active students: ***')
time.sleep(0.5)
print(
    'Added Registration Status' + "\nstudent's advisor" +
    '\nAdded campusIDs' + "\nAdded student's majors" \
    + '\nAdded Unit registration notes')
