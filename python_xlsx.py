import os
os.getcwd()
import openpyxl


wb = openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames)

sheet = wb.sheetnames

ws = wb.active
print(ws)

cell = ws['A1'].value
print(cell)

cell_range = ws['A1':'C2']
print(cell_range)

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell.value)


x = 'Row ' + str(cell.row) + ', Column ' + cell.column + ' is ' + str(cell.value)
print(x)

cell = ws.cell(row=1, column=2).value
print(cell)

for i in range(1, 8, 2):
    print(i, ws.cell(row=i, column=2).value)


sheet = ws['A1':'C3']

print('\n')
for rowOfCellObjects in sheet:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('---END OF ROW---')


for cell in ws['B']:
   print(cell.value)

