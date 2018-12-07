import openpyxl


file1 = '/Users/nbenzschawel/Desktop/New_sevis_data.xlsx'
file2 = '/Users/nbenzschawel/Desktop/SEVIS_transfer_data.xlsx'

wb = openpyxl.load_workbook(file1)
wb2 = openpyxl.load_workbook(file2)
old_sheet = wb.worksheets[0]
new_sheet = wb2.worksheets[1]
final_sheet = wb2.worksheets[0]

row_max_old = old_sheet.max_row
col_max_old = old_sheet.max_column

row_max_current = new_sheet.max_row
col_max_current = new_sheet.max_column

row_max_final = final_sheet.max_row
col_max_final = final_sheet.max_column


# pasteRange(1, new_sheet.max_row+1, new_sheet.max_column, rowmax_new, new_sheet, selectedRange)