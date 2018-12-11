import openpyxl
from SEVIS_Transfers.build_new_student_data import file2, \
    wb2, old_sheet, new_sheet, row_max_old, col_max_old, final_sheet, \
    row_max_current, col_max_current, row_max_final, col_max_final

from SEVIS_Transfers.sort_data import sort_data
ws2 = wb2.active


def copy_new_Range(startCol, startRow, endCol, endRow, old_sheet):
    """
    Copies a range of cells as a nested list
    Takes: start cell, end cell, and the sheet we want to copy from
    """
    rangeSelected = []
    for i in range(startRow, endRow+1, 1):
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(old_sheet.cell(row = i, column = j).value)
        rangeSelected.append(rowSelected)

    return rangeSelected


def paste_new_Range(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    """Pastes the selected data from copy_new_Range to your desired output worksheet"""
    countRow = 0
    for i in range(startRow, endRow+1, 1):
        countCol = 0
        for j in range(startCol, endCol+1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1


def create_new_Data():
    """Pastes copied range to new sheet which needs to be updates"""
    print(f'Old Row Range = {new_sheet.max_row}')
    print('Processing data...')

    if new_sheet.cell(row=1, column=1).value is None:
        selectedRange = copy_new_Range(1, 1, col_max_old, row_max_old, old_sheet)
        paste_new_Range(1, 1, col_max_old, row_max_old, new_sheet, selectedRange)

    elif new_sheet.cell(row=1, column=1).value == 'SEVIS ID':
        selectedRange = copy_new_Range(1, 2, col_max_old, row_max_old, old_sheet)
        paste_new_Range(1, (row_max_current+1), col_max_old, (row_max_current + row_max_old-1), new_sheet, selectedRange)

    wb2.save(file2)
    print("Range copied and pasted")
    print(f'New Row Range = {new_sheet.max_row}')


def paste_to_final():
    """Pastes copied range from new sheet to final sheet - do this after sorting data"""
    print(f'Current Row Range = {row_max_current}')
    print('Processing data...')
    selectedRange = copy_new_Range(1, 1, col_max_current, row_max_current, new_sheet)
    paste_new_Range(1, 1, col_max_current, row_max_current, final_sheet, selectedRange)
    wb2.save(file2)
    print("Range copied and pasted")
    print(f'New Row Range = {row_max_final}')


def grab_final_data():
    """Copies old "Final" data into new_sheet to allow updated data comparison."""
    print(f'Current Row Range = {final_sheet.max_row}')
    print('Grabbing most recent data...')
    selectedRange = copy_new_Range(1, 1, col_max_final, row_max_final, final_sheet)
    paste_new_Range(1, 1, col_max_final, row_max_final, new_sheet, selectedRange)
    wb2.save(file2)
    print('Range copied and pasted')
    print(f'New Row Range = {new_sheet.max_row}')


def check_in_fsa():
    """Marks new data that needs to be input into ISSM"""
    title = new_sheet.cell(row=1, column=10)
    title.value = 'ISSM Status'
    for rowNum in range(1, row_max_current):
        status = new_sheet.cell(row=rowNum, column=9).value
        if status == 'INITIAL':
            new_sheet.cell(row=rowNum, column=10).value = 'Add to ISSM'
    print('Added notes for ISSM')

    wb2.save(file2)


def find_in_fsa():
    """Checks the status of final data to see if it is already in ISSM"""
    for rowNum in range(1, row_max_final):
        status = final_sheet.cell(row=rowNum, column=10).value
        if status == 'Add to ISSM':
            final_sheet.cell(row=rowNum, column=10).value = 'Old Data: in ISSM'
    print('Updated ISSM notes for new data')

    wb2.save(file2)


paste_to_final()