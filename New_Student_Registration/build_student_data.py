import os
os.getcwd()
import openpyxl
from Handlers.file_imports import wb1_new, wb2_new, ws1_new, ws2_new

"""
This file builds dictionaries containing the following data: 
campusIDs : SEVISIDs
SEVISIDs : Majors

"""


campusID = []
SEVISID = []

for rowNum in range(2, ws1_new.max_row):  # skip the first row and go to the last row
    id = ws1_new.cell(row=rowNum, column=5).value
    SEVISID.append(id)

for rowNum in range(2, ws1_new.max_row):
    id = ws1_new.cell(row=rowNum, column=1).value
    campusID.append(id)

# build dict where k = SEVISID, v = campusID
campusID_SEVISID = {}

for i in range(len(SEVISID)):
    campusID_SEVISID[SEVISID[i]] = campusID[i]

# build a list of majors
major_data = []

for rowNum in range(2, ws1_new.max_row):
    major = ws1_new.cell(row=rowNum, column=13).value
    major_data.append(major)

# build a dict where k = SEVISID, v = list of majors
SEVISID_major = dict(zip(SEVISID, major_data))









