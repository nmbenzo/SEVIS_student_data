import os
os.getcwd()
import openpyxl


wb2 = openpyxl.load_workbook('SEVIS_raw_data.xlsx')

sheet_builder = wb2.sheetnames


ws2 = wb2.active


SEVIS_look_for = []

for rowNum in range(2, ws2.max_row):  # skip the first row and go to the last row
    id = ws2.cell(row=rowNum, column=4).value
    SEVIS_look_for.append(id)

print(SEVIS_look_for)