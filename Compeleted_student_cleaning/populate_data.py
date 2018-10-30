import os
os.getcwd()
import openpyxl
from Compeleted_student_cleaning.build_comp_data import SEVISID

wb2 = openpyxl.load_workbook('/Users/nbenzschawel/Downloads/Completed_Students_Cleanup_Report_201840.xlsx')

sheet = wb2.worksheets[0]

ws = wb2.active


def match_SEVISID(SEVISID):
    """
    match the SEVISID from the SEVIS spreadsheet of completed students with the
    SEVISID from ACTIVE status students report out of Atlas. If there is an ID
    match, populate new column with notes to close.
    """
    sheet.insert_cols(1)
    title = sheet.cell(row=1, column=1)
    title.value = 'Completed Notes'
    for row in range(2, sheet.max_row):
        sevis_ID = sheet.cell(row=row, column=3).value
        for x in SEVISID:
            if x == sevis_ID:
                sheet.cell(row=row, column=1). value = 'Student is completed in SEVIS, but ACTIVE in ISSM.'

    wb2.save('/Users/nbenzschawel/Downloads/Completed_Students_Cleanup_Report_201840.xlsx')


match_SEVISID(SEVISID)