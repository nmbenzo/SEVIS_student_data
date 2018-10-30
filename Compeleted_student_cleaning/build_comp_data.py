import os
os.getcwd()
import openpyxl

wb1 = openpyxl.load_workbook('/Users/nbenzschawel/Downloads/SEVIS - Completed Status Students (past 18 months).xlsx')

ws1 = wb1.active


SEVISID = []

for row in range(2, ws1.max_row):
    id = ws1.cell(row=row, column=1).value
    SEVISID.append(id)




