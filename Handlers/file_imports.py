import openpyxl
from Handlers.directories import path_to_raw_data, path_to_final_data_F19, \
path_to_final_data_S20, main_path, location_a_path, location_b_path


"""Master Files"""
try:
		Registration_file = main_path + '2020/Spring/SEVIS Registration Spring 2020.xlsx'

		SEVIS_Live_Workbook = main_path + '2020/Spring/SEVIS Registration Live.xlsx'


		"""Master Files"""
		major_advisor_file = main_path + '2019/Fall/Raw_Files/major_advisor.xlsx'


		""" File Location for Google_Drive.drive_main.py"""

		location_a = location_a_path + 'SEVIS_Application/SEVIS Registration.xlsx'
		location_b = location_b_path + \
					 'SEVIS_Reg/2020/Spring/SEVIS Registration Live.xlsx'


		""" COL Student Data """

		COL_students_raw = path_to_raw_data + \
						   '2019/Fall/Raw_Files/REPORT_STUDENT_STATUS_COL.xlsx'

		COL_students_FINAL = path_to_final_data_F19 + 'COL_Final.xlsx'

		COL_wb = openpyxl.Workbook()

		try:
			COL_students = openpyxl.load_workbook(COL_students_FINAL)
			COL_students_sheet = COL_students.worksheets[0]
		except FileNotFoundError:
			COL_wb.save(COL_students_FINAL)
			COL_students = openpyxl.load_workbook(COL_students_FINAL)
			COL_students_sheet = COL_students.worksheets[0]


		""" New Student Data """

		# ISSM Report: SEVIS Registration Tracking-New Students-SEVIS Pending
		new_stud_issm_data = path_to_raw_data + \
		'2020/Spring/Raw_Files/SEVIS_Registration_Tracking-New_Students-SEVIS_Pending.csv'

		# SEVIS Report: Initial Status Students
		sevis_inital_student_data = path_to_raw_data +\
						'2020/Spring/Raw_Files/SEVIS - Initial Status Students.csv'


		# Need to sort Z-A by Event Status Column to ensure Submitted is matched before
		# Archived

		issm_add_address_phone = path_to_raw_data +\
		'2019/Fall/Raw_Files/SEVIS_Registration_Tracking-New_Students-SEVIS_Pending_09112019.xlsx'

		issm_reg_data_success = path_to_raw_data +\
		'2019/Fall/Raw_Files/All_SEVIS-Active_Student_Tracking_' \
		'(Registration_Events_Returned).xlsx'

		NEW_students_FINAL = path_to_final_data_S20 + 'NEW_Final.xlsx'

		NEW_final_MASTER = path_to_final_data_F19 + 'NEW_FINAL_MASTER.xlsx'

		NEW_wb = openpyxl.Workbook()

		try:
			NEW_students = openpyxl.load_workbook(NEW_students_FINAL)
			NEW_students_sheet = NEW_students.worksheets[0]
		except FileNotFoundError:
			NEW_wb.save(NEW_students_FINAL)
			NEW_students = openpyxl.load_workbook(NEW_students_FINAL)
			NEW_students_sheet = NEW_students.worksheets[0]

		# wb1_new = openpyxl.load_workbook(new_stud_issm_data)  # workbook from ISSM
		# wb2_new = openpyxl.load_workbook(
		#	sevis_inital_student_data)  # workbook from SEVIS RTI

		# ws1_new = wb1_new.active
		# ws2_new = wb2_new.active


		""" SEVIS Active Student Data """

		active_stud_issm_data = path_to_raw_data + \
						'2020/Spring/Raw_Files/All_SEVIS-Active_Student_Tracking.csv'

		active_student_req_reg = path_to_raw_data + \
			'2020/Spring/Raw_Files/SEVIS - Active Students Requiring Registration.csv'

		active_stud_issm_data_F19 = path_to_raw_data + \
						'2019/Fall/Raw_Files/All_SEVIS-Active_Student_Tracking.xlsx'


		active_student_req_reg_f = path_to_raw_data + \
			'2019/Fall/Raw_Files/SEVIS - Active Students Requiring Registration_final.xlsx'

		active_stud_reg_event_status = path_to_raw_data + ""


		ACTIVE_students_FINAL = path_to_final_data_S20 + 'ACTIVE_Final.xlsx'


		ACTIVE_wb = openpyxl.Workbook()

		try:
			ACTIVE_students = openpyxl.load_workbook(ACTIVE_students_FINAL)
			ACTIVE_students_sheet = ACTIVE_students.worksheets[0]
		except FileNotFoundError:
			ACTIVE_wb.save(ACTIVE_students_FINAL)
			ACTIVE_students = openpyxl.load_workbook(ACTIVE_students_FINAL)
			ACTIVE_students_sheet = ACTIVE_students.worksheets[0]

		# wb2_active = openpyxl.load_workbook(active_student_req_reg)
		# sheet = wb2_active.worksheets[0]
		# ws2 = wb2_active.active
		#
		# wb1 = openpyxl.load_workbook(active_stud_issm_data)
		# ws = wb1.active


		""" Transfer Student Data"""

		new_transfer_data = path_to_raw_data + \
						'2019/Spring/Raw_Files/SEVIS - Students Transferring In.xlsx'

		current_transfer_data = path_to_raw_data + \
								'2019/Spring/Raw_Files/SEVIS_transfer_data.xlsx'

		wb_trans = openpyxl.load_workbook(new_transfer_data)
		wb2_trans = openpyxl.load_workbook(current_transfer_data)
		old_sheet = wb_trans.worksheets[0]
		new_sheet = wb2_trans.worksheets[1]
		final_sheet = wb2_trans.worksheets[0]

		row_max_old = old_sheet.max_row
		col_max_old = old_sheet.max_column

		row_max_current = new_sheet.max_row
		col_max_current = new_sheet.max_column

		row_max_final = final_sheet.max_row
		col_max_final = final_sheet.max_column

		transfer_stud_issm_data_S20 = path_to_raw_data + \
						'2020/Spring/Raw_Files/All_SEVIS-Pending_Student_Tracking.csv'

		transfer_stud_sevis_data_S20 = path_to_raw_data + \
						'2020/Spring/Raw_Files/SEVIS - Students Transferring In.csv'

		TRANS_students_FINAL = path_to_final_data_S20 + 'TRANS_Final.xlsx'


		TRANS_wb = openpyxl.Workbook()

		try:
			TRANS_students = openpyxl.load_workbook(TRANS_students_FINAL)
			TRANS_students_sheet = TRANS_students.worksheets[0]
		except FileNotFoundError:
			TRANS_wb.save(TRANS_students_FINAL)
			TRANS_students = openpyxl.load_workbook(TRANS_students_FINAL)
			TRANS_students_sheet = TRANS_students.worksheets[0]



		""" No Show-Cancels for Admissions"""

		SEVIS_initial_status_stud = path_to_raw_data + \
						'2020/Spring/Raw_Files/SEVIS - Initial Status Students.xlsx'

		No_show_UG = path_to_raw_data + '2020/Spring/Raw_Files/No-Shows-UG.xlsx'

		No_show_GR = path_to_raw_data + '2020/Spring/Raw_Files/No-Shows-Grad.xlsx'

		No_SHOW_students_FINAL = path_to_final_data_S20 + 'NO_SHOW_Final.xlsx'

		No_SHOW_wb = openpyxl.Workbook()

		try:
			NO_SHOW_students = openpyxl.load_workbook(No_SHOW_students_FINAL)
			NOSHOW_students_sheet = NO_SHOW_students.worksheets[0]
		except FileNotFoundError:
			No_SHOW_wb.save(No_SHOW_students_FINAL)
			NO_SHOW_students = openpyxl.load_workbook(No_SHOW_students_FINAL)
			NOSHOW_students_sheet = NO_SHOW_students.worksheets[0]

		# SEV_initial_data = openpyxl.load_workbook(SEVIS_initial_status_stud)
		# wb_cancel_ug = openpyxl.load_workbook(No_show_UG)
		# wb_cancel_gr = openpyxl.load_workbook(No_show_GR)
		#
		# sevis_initial = SEV_initial_data.worksheets[0]
		# ug_sheet = wb_cancel_ug.worksheets[0]
		# gr_sheet = wb_cancel_gr.worksheets[0]
		#
		# initial_max_row = sevis_initial.max_row
		# ug_row_max = ug_sheet.max_row
		# ug_col_max = ug_sheet.max_column
		# gr_row_max = gr_sheet.max_row
		# gr_col_max = gr_sheet.max_column
		#
		# sevis_initial_ws = SEV_initial_data.active
		# ws_cancel_ug = wb_cancel_ug.active
		# ws_cancel_gr = wb_cancel_gr.active


		""" Completed Student Data """

		comp_student_cleanup_report = path_to_raw_data + \
				'2019/Spring/Raw_Files/Completed_Students_Cleanup_Report_201840.xlsx'

		sevis_completed_students = path_to_raw_data + \
		'2019/Spring/Raw_Files/SEVIS - Completed Status Students (past 18 months).xlsx'

		completed_student_wb2 = openpyxl.load_workbook(comp_student_cleanup_report)

		comp_sheet = completed_student_wb2.worksheets[0]

		comp_ws = completed_student_wb2.active

		wb1_complete = openpyxl.load_workbook(sevis_completed_students)

		ws1 = wb1_complete.active


		""" Graduated Student Data """

		# graduated_students = path_to_raw_data + \
		# 				'2019/Fall/Raw_Files/graduated_students_201920_201930.xlsx'
		# wb1_grad = openpyxl.load_workbook(active_stud_issm_data)
		#
		# sheet_grad = wb1_grad.worksheets[0]
		# ws_grad = wb1_grad.active


		# wb2_grad = openpyxl.load_workbook(path_to_raw_data +
		# 					'2019/Fall/Raw_Files/graduated_students_201920_201930.xlsx')
		# wb3Live_grad = openpyxl.load_workbook(graduated_students)
		#
		# sheet2_grad = wb2_grad.worksheets[0]
		# sheet3_grad = wb3Live_grad.worksheets[0]
		#
		# ws2_grad = wb2_grad.active
		# ws3_grad = wb3Live_grad.active


		"""SEVIS Registration Timelines"""

		current_registration_timeline = \
		'/Volumes/ISS/Staff Only/Student Lists/SEVIS Registration/' \
		'SEVIS Registration Timelines/SEVIS REGISTRATION S20.docx'


		"""J-1 Students"""

		J_students_FINAL = path_to_raw_data + \
						'2019/Spring/Raw_Files/SEVIS_Validation_Tracking - New_EVs.xlsx'
		J_wb = openpyxl.Workbook()

		try:
			J_students = openpyxl.load_workbook(J_students_FINAL)
			J_students_sheet = J_students.worksheets[0]
		except FileNotFoundError:
			J_wb.save(J_students_FINAL)
			J_students = openpyxl.load_workbook(J_students_FINAL)
			J_students_sheet = J_students.worksheets[0]
except:
	pass

