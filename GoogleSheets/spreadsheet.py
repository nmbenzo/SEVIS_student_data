import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import date, datetime
import openpyxl


def authenticate_google_docs():
    # use creds to create a client to interact with the Google Drive API
    scope = ['https://spreadsheets.google.com/feeds',
             'https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_name('client_secret.json', scope)
    gc = gspread.authorize(creds)
    return gc

wb2 = openpyxl.load_workbook('/Users/nbenzschawel/Downloads/SEVIS_raw_data.xlsx') # workbook from SEVIS RTI
ws2 = wb2.active


gc = authenticate_google_docs()

sh = gc.open('201920 SEVIS Registration')
worksheet_list = sh.worksheets()


# Find a workbook by name and open a particular sheet
# Make sure you use the right name here.

COL = sh.worksheet('COL')
transferUG = sh.worksheet('Transfer UG')
transferGR = sh.worksheet('Transfer GR')
new_students = sh.worksheet('New Students')
continuing_students = sh.worksheet('Continuing Students')


def update_new_students(ws2):
    max_row = ws2.max_row
    max_column = ws2.max_column
    for rowNum in range(1, max_row+1):
        for colNum in range(1, max_column+1):
            cell_value = ws2.cell(row=rowNum, column=colNum).value
            if isinstance(cell_value, (datetime, date)):
                new_students.update_cell(row=rowNum, col=colNum, value=cell_value.isoformat())
            else:
                new_students.update_cell(row=rowNum, col=colNum, value=cell_value)

update_new_students(ws2)
