import os
os.getcwd()
import openpyxl
from major_advisor_data import advisor_major_ug, advisor_major_gr

wb = openpyxl.load_workbook('SEVIS_Registration_Tracking-New_Students-SEVIS_Pending_09072018094611.xlsx')
print(wb.sheetnames)

sheet = wb.sheetnames

ws = wb.active
print(ws)

for row in ws.iter_rows(min_row=2, max_col=6, max_row=5):
    for cell in row:
        print(cell.value)


sheet = ws['A1':'F39']

print('\n')
for rowOfCellObjects in sheet:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('---END OF ROW---')


for i in range(1, 8, 1):
    print(i, ws.cell(row=i, column=6).value)

print('\n')

def match_advisor(advisor_major_ug, advisor_major_gr):
    for rowNum in range(2, ws.max_row): # skip the first row and go to the last row
     major = ws.cell(row=rowNum, column=11).value
     if major in advisor_major_ug:
         ws.cell(row=rowNum, column=9).value = advisor_major_ug[major]
     if major in advisor_major_gr:
         ws.cell(row=rowNum, column=9).value = advisor_major_gr[major]



